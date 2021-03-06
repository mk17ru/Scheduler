from openpyxl import load_workbook
import re
import pyexcel as p

p.save_book_as(file_name='./data/Приложение №5.xls',
               dest_file_name='./data/Приложение №5.xlsx')

def parsing_week():
    wb = load_workbook('data/Приложение №1.xlsx')
    sheet = wb.get_sheet_by_name('ДПО')
    week_amount = [0, 5, 4, 4, 5, 4, 4, 5, 4, 4, 5, 4, 5]
    weeks = []
    hidden = {"Начальная подготовка СД", "DCS Астра"}
    discipline_col = 7
    year_col = 9
    for month in range(12):
        year_col = year_col + week_amount[month] * 2
        for week in range(year_col, year_col + week_amount[month + 1] * 2 + 1, 2):
            for p in range(6, 30, 2):
                if sheet.cell(row=p, column=discipline_col).value == None or sheet.cell(row=p, column=week).value == None:
                    continue
                else:
                    s = sheet.cell(row=p, column=week).value.split('\n')
                    while len(s) > 0:
                        if (len(s) > 3):
                            new_str = ""
                            ind = 0
                            for r in s:
                                if len(r) > 0 and str.isalpha(r[0]):
                                    new_str += r + " "
                                    ind += 1
                                else:
                                    break
                            new_s = []
                            new_s.append(new_str.strip())
                            for i in range(ind, ind + 2):
                                new_s.append(s[i])
                            s_cop = []
                            for z in range(ind + 2, len(s)):
                                if (s[z] != ""):
                                    s_cop.append(s[z])
                            s = s_cop
                        else:
                            new_s = s
                            s = []
                        # because of hidden cells
                        if (len(new_s) < 3 or new_s[0] == "" or new_s[1] == "" or new_s[2] == ""):
                            continue
                        auditoriums = []
                        new_s[2] = str.replace(new_s[2], '.', ' ')
                        for i in re.split(' ', new_s[2]):
                            if i == "ауд" or i == "ИАС" or i == "":
                                continue
                            for cur in i.split(','):
                                auditoriums.append(cur)
                        new_s[1] = new_s[1].replace(" ",  '')
                        new_s[1] = new_s[1].replace("–", "-")
                        if not(hidden.__contains__(sheet.cell(row=p, column=discipline_col).value)):
                            temp = {"discipline": sheet.cell(row=p, column=discipline_col).value, "type": new_s[0], "date": new_s[1], "classroom": auditoriums}
                            weeks.append(temp)
    return weeks

matching_disp = {"Программа повышения квалификации \"Центровка и контроль загрузки воздушных судов. Базовый курс\"" : {"discipline": "Центровка и контроль загрузки ВС", "type" : "Базовый курс"},
                 "Программа повышения квалификации \"Центровка и контроль загрузки воздушных судов\"" : {"discipline" : "Центровка и контроль загрузки ВС", "type" : "Повышения квалификации"},
                "Программа повышения квалификации «Перевозка опасных грузов воздушным транспортом.8 категория ИКАО/ИАТА\"" : {"discipline" : "Опасные грузы. 8 категория", "type" : "Программа повышения квалификации"},
                "Программа повышения квалификации «Перевозка опасных грузов воздушным транспортом. 8 категория ИКАО/ИАТА.Базовый курс»" : {"discipline" : "Опасные грузы. 8 категория", "type" : "Базовый курс"},
                 "Программа повышения квалификации «Перевозка опасных грузов воздушным транспортом.9 категория ИКАО/ИАТА\"": {"discipline": "Опасные грузы. 9 категория", "type": "Программа повышения квалификации"},
                 "Программа повышения квалификации «Перевозка опасных грузов воздушным транспортом. 9 категория ИКАО/ИАТА.Базовый курс»": {"discipline": "Опасные грузы. 9 категория", "type": "Базовый курс"},
                 "Программа повышения квалификации «Перевозка опасных грузов воздушным транспортом. 10 категория ИКАО/ИАТА\"": {"discipline": "Опасные грузы. 10 категория", "type": "Программа повышения квалификации"},
                 "Программа повышения квалификации «Перевозка опасных грузов воздушным транспортом. 10 категория ИКАО/ИАТА.Базовый курс»": {"discipline": "Опасные грузы. 10 категория", "type": "Базовый курс"},
                 "Программа повышения квалификации «Организация обслуживания пассажирских перевозок воздушным транспортом»" : {"discipline": "Пассажирские перевозки", "type": "Программа повышения квалификации"},
                 "Программа повышения квалификации «Организация обслуживания пассажирских перевозок воздушным транспортом.Базовый курс»" : {"discipline": "Пассажирские перевозки", "type": "Базовый курс"},
                 "Программа повышения квалификации «Организация наземного обслуживания воздушных судов. Базовый курс»" : {"discipline": "Организация наземного обслуживания ВС", "type": "Базовый курс"},
                 "Программа повышения квалификации «Организация наземного обслуживания воздушных судов»" : {"discipline": "Организация наземного обслуживания ВС", "type": "Программа повышения квалификации"},
                "Программа специальной профессиональной подготовки «Предполётный досмотр пассажиров, членов экипажей гражданских судов, обслуживающего персонала, ручной клади, багажа, грузов, почты и бортовых запасов»" : {"discipline": "Досмотр", "type": "Специальная профессиональная подготовка"},
                "Программа повышения квалификации «Предполётный досмотр пассажиров, членов экипажей гражданских судов, обслуживающего персонала, ручной клади, багажа, грузов, почты и бортовых запасов»" : {"discipline": "Досмотр", "type": "Программа повышения квалификации"},
                "Программа специальной профессиональной подготовки «Перронный контроль и досмотр воздушных судов»": {"discipline": "Перронный контроль", "type": "Специальная профессиональная подготовка"},
                 "Программа повышения квалификации «Перронный контроль и досмотр воздушных судов»": {"discipline": "Перронный контроль", "type": "Программа повышения квалификации"},
                 "Программа специальной профессиональной подготовки  «Предотвращение несанкционированного доступа в контролируемую зону аэропорта»" : {"discipline": "Охрана аэропорта", "type": "Специальная профессиональная подготовка"},
                 "Программа повышения квалификации «Предотвращение несанкционированного доступа в контролируемую зону аэропорта»" : {"discipline": "Охрана аэропорта", "type": "Программа повышения квалификации"}
                 }


teachers_holidays = {}

def parse_teachers():
    wb = load_workbook('data/Приложение №2.xlsx')
    sheet3 = wb.get_sheet_by_name('параметры преподавателей')
    sheet1 = wb.get_sheet_by_name('параметры программ')
    teachers = []
    for p in range(2, 35):
        st = list(map(int, str(sheet3.cell(row=p, column=4).value).split(';')))
        disc = []
        for i in st:
            if matching_disp.__contains__(sheet1.cell(row=i + 1, column=3).value.strip()):
                disc.append(matching_disp[sheet1.cell(row=i + 1, column=3).value.strip()])
        if len(disc) > 0:
            teachers.append({"name" : sheet3.cell(row=p, column=2).value, "works" : disc, "prior" : sheet3.cell(row=p, column=5).value,
                        "work_time" : sheet3.cell(row=p, column=7).value, "scheduler" : sheet3.cell(row=p, column=8).value, "holidays": teachers_holidays.get(sheet3.cell(row=p, column=2).value)
                             })
    return teachers

def parse_disc():
    wb = load_workbook('data/Приложение №2.xlsx')
    sheet1 = wb.get_sheet_by_name('параметры программ')
    new_d = []
    for p in range(2, 41):
        if matching_disp.__contains__(sheet1.cell(row=p, column=3).value.strip()):
            temp = matching_disp[sheet1.cell(row=p, column=3).value.strip()]
            temp.update({"hours": sheet1.cell(row=p, column=5).value})
            new_d.append(temp)
    return new_d

def parse_holidays():
    wb = load_workbook('data/Приложение №5.xlsx')
    sheet1 = wb.get_sheet_by_name('План')
    for p in range(9, 24):
        for i in range(4, 27, 2):
            if sheet1.cell(row=p, column=i).value != None:
                cur_data = ""
                cur_data += str((sheet1.cell(row=p, column=i + 1).value-1) * 10 + 1) + "." + str((i - 2) // 2) + "-"
                if (sheet1.cell(row=p, column=i + 1).value-1) * 10 + sheet1.cell(row=p, column=i).value > 30:
                    cur_data += str(sheet1.cell(row=p, column=i).value - (30 - (sheet1.cell(row=p, column=i + 1).value-1) * 10)) + "." + str(((i - 2) // 2 + 1) % 13)
                else:
                    cur_data += str((sheet1.cell(row=p, column=i + 1).value-1) * 10 + sheet1.cell(row=p, column=i).value) + "." + str((i - 2) // 2)

                if not(teachers_holidays.__contains__(sheet1.cell(row=p, column=2).value)):
                    teachers_holidays[sheet1.cell(row=p, column=2).value] = []
                teachers_holidays[sheet1.cell(row=p, column=2).value].append(cur_data)


                #Parsing application 1
for i in parsing_week():
    print(i, end=", \n")
print()
print("------------------------------------------------------------\n upper were weeks")

#Parsing application 5
parse_holidays()

#Parsing application 2
for j in parse_teachers():
    print(j, end=", \n")
print()
print("------------------------------------------------------------\n upper were teachers")
for z in parse_disc():
    print(z, end=", \n")
print()
print("------------------------------------------------------------\n upper were hours")

print(teachers_holidays)
#parse_teachers()